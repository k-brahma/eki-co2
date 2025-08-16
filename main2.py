import os
import requests
from dotenv import load_dotenv
import openpyxl

load_dotenv()
api_key = os.getenv('EKISPERT_API_KEY')

base_url = 'https://api.ekispert.jp/v1/json/search/course/plain?key={API_KEY}&from={station_from}&to={station_to}'

# Excelファイルを開く
wb = openpyxl.load_workbook('通勤経路.xlsx')
ws = wb.active

# 2-11行のデータを処理
for row in range(2, 12):
    station_from = ws[f'D{row}'].value  # D列：最寄り駅
    station_to = ws[f'E{row}'].value  # E列：出勤先駅

    print(row, station_from, station_to)

    # APIリクエスト
    url = base_url.format(API_KEY=api_key, station_from=station_from, station_to=station_to)
    response = requests.get(url)
    data = response.json()

    # 最初の経路を取得
    course = data['ResultSet']['Course'][0]
    co2_emission = course['Route']['exhaustCO2']

    # F列にCO2排出量を書き込み
    ws[f'F{row}'] = int(co2_emission)

    print(f"{station_from} → {station_to}: {co2_emission}g")

# Excelファイルを保存
wb.save('通勤経路.xlsx')
print("CO2排出量をExcelファイルに保存しました！")